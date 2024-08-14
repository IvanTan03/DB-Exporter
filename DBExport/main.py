from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivymd.app import MDApp
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDRaisedButton
from kivy.properties import ObjectProperty
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.popup import Popup
from kivymd.uix.pickers import MDDatePicker
from kivy.clock import Clock
import threading
from kivy.uix.textinput import TextInput

import pandas as pd
import mysql.connector
from mysql.connector import Error
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import load_workbook
from openpyxl.styles import Alignment

class ReportExportWidget(BoxLayout):
    from_date = ObjectProperty(None)
    to_date = ObjectProperty(None)
    pdf_checkbox = ObjectProperty(None)
    excel_checkbox = ObjectProperty(None)
    status_label = ObjectProperty(None)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.orientation = 'vertical'
        self.padding = 20
        self.spacing = 20
        self.load_stylesheet()
        Clock.schedule_once(self.check_connection, 0)

    def load_stylesheet(self):
        pass

    def show_date_picker(self, date_field):
        date_dialog = MDDatePicker()
        date_dialog.bind(on_save=lambda instance, value, date_range: self.set_date(date_field, value))
        date_dialog.open()

    def set_date(self, date_field, date_value):
        date_field.text = str(date_value)

    def export_report(self):
        from_date = self.from_date.text
        to_date = self.to_date.text
        pdf_selected = self.pdf_checkbox.state == 'down'
        excel_selected = self.excel_checkbox.state == 'down'

        if not from_date or not to_date:
            self.show_dialog("Error", "Please select both from and to dates.")
            return

        if not pdf_selected and not excel_selected:
            self.show_dialog("Error", "No export format selected.")
            return

        try:
            data = self.fetch_data(from_date, to_date)
            if data is None:
                raise Exception("Failed to fetch data.")

            file_chooser = FileChooserIconView()
            file_chooser.path = ''  # Start with an empty path
            file_chooser.filters = ["*"]
            confirm_button = MDRaisedButton(
                text="Select Directory",
                size_hint_y=None,
                height='40dp',
                md_bg_color=[0.1, 0.5, 0.5, 1],
                text_color=[1, 1, 1, 1],
                on_release=lambda *x: self.show_filename_input(file_chooser.path, data, pdf_selected, excel_selected, from_date, to_date)
            )

        # Create a layout for the popup content
            content = BoxLayout(orientation='vertical')
            content.add_widget(file_chooser)
            content.add_widget(confirm_button)
            popup = Popup(title="Select Save Directory", content=content, size_hint=(0.8, 0.8))
            popup.open()

        except Exception as e:
            self.show_dialog("Error", f"Export failed: {e}")

    def show_filename_input(self, directory, data, pdf_selected, excel_selected, from_date, to_date):
        if not directory:
            self.show_dialog("Export Cancelled", "No directory chosen.")
            return

        content = BoxLayout(orientation='vertical', spacing=10, padding=[20, 20, 20, 20])

        filename_input = TextInput(hint_text="Enter filename", size_hint_y=None, height='40dp', padding=[10, 5])
        content.add_widget(filename_input)

        spacer = BoxLayout(size_hint_y=None, height='10dp')
        content.add_widget(spacer)

        confirm_button = MDRaisedButton(
            text="Save",
            size_hint_y=None,
            height='40dp',
            md_bg_color=[0.1, 0.5, 0.5, 1],
            text_color=[1, 1, 1, 1],
            on_release=lambda *x: self.save_file(directory, filename_input.text.strip(), data, pdf_selected, excel_selected, from_date, to_date)
        )
        content.add_widget(confirm_button)

    # Create and open the popup
        popup = Popup(
            title="Enter Filename",
            content=content,
            size_hint=(0.8, 0.4)
        )
        popup.open()


    def save_file(self, directory, filename, data, pdf_selected, excel_selected, from_date, to_date):
        
        if not directory:
            self.show_dialog("Export Cancelled", "No directory chosen.")
            return
        
        if not filename:
            self.show_dialog("Error", "Please enter a filename.")
            return

    # Ensure the directory ends with a '/'
        directory = directory.rstrip('/') + '/'
        file_path = f"{directory}{filename}"

        if pdf_selected:
            file_path += '.pdf'
            self.convert_to_pdf(data, file_path, from_date, to_date)
        elif excel_selected:
            file_path += '.xlsx'
            self.convert_to_excel(data, file_path)

        self.show_dialog("Success", "Export successful.")


    def fetch_data(self, from_date, to_date):
        try:
            connection = mysql.connector.connect(
                host='192.168.1.220',
                database='ztest',
                user='test',
                password='123',
                charset='utf8'
            )

            if connection.is_connected():
                query = f"""
                SELECT
                H.D_ate,
                H.Doc1No,
                H.HCNetAmt,
                H.HCDtTax,
                D.CashAmt,
                D.CC1Code,
                D.CC1Amt,
                D.CC2Code,
                D.CC2Amt,
                D.Cheque1Amt,
                D.Cheque2Amt,
                D.BalanceAmount,
                D.ChangeAmt,
                D.CounterCode
                FROM
                stk_cus_inv_hd AS H
                INNER JOIN stk_receipt2 AS D ON H.Doc1No = D.Doc1No
                WHERE
                H.D_ate BETWEEN '{from_date}' AND '{to_date}'
                """
                data = pd.read_sql(query, connection)
                connection.close()
                return data
            else:
                self.show_dialog("Error", "Error connecting to database.")
                return None
        except Error as e:
            self.show_dialog("Error", f"Database error: {e}")
            return None
        except Exception as e:
            self.show_dialog("Error", f"Unexpected error: {e}")
            return None

    def convert_to_pdf(self, data, filename, from_date, to_date):
        left_margin = 20
        right_margin = 20
        top_margin = 20
        bottom_margin = 20

        doc = SimpleDocTemplate(
            filename,
            pagesize=landscape(letter),
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=top_margin,
            bottomMargin=bottom_margin
        )

        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_text = f"Report from {from_date} to {to_date}"
        title = Paragraph(title_text, title_style)
        elements.append(title)

        # Add a column with row numbers
        data.insert(0, "No.", range(1, len(data) + 1))
        data_list = [data.columns.tolist()] + data.values.tolist()

        min_col_width = 50
        max_col_width = 150
        column_widths = []

        for col_index in range(len(data.columns)):
            max_length = max(data[data.columns[col_index]].astype(str).apply(len).max(), len(data.columns[col_index]))
            col_width = min(max_length * 9, max_col_width)
            column_widths.append(max(min_col_width, col_width))

        total_width = sum(column_widths)
        page_width = landscape(letter)[0] - (left_margin + right_margin)
        if total_width > page_width:
            scaling_factor = page_width / total_width
            column_widths = [w * scaling_factor for w in column_widths]

        header_style = styles['Normal']
        header_style.fontName = 'Helvetica-Bold'
        header_style.fontSize = 10
        header_style.textColor = colors.white
        header_style.wordWrap = 'CJK'

        wrapped_headers = [Paragraph(f"<para>{header}</para>", header_style) for header in data.columns]
        data_list[0] = wrapped_headers

        table = Table(data_list, colWidths=column_widths)

        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ])
        table.setStyle(table_style)

        elements.append(table)
        doc.build(elements)

    def convert_to_excel(self, data, filename):
        data.to_excel(filename, index=False)

        workbook = load_workbook(filename)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value is not None:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')

        workbook.save(filename)
        print(f"Converting data to Excel and saving as {filename}")

    def show_dialog(self, title, message):
        dialog = MDDialog(title=title, text=message, size_hint=(0.8, 0.4))
        dialog.open()

    def check_connection(self, *_):
        def update_status_label(text, color):
            if self.status_label:
                print(f"Updating status label to: {text}")  # Debugging line
                self.status_label.text = text
                self.status_label.color = color
        
        def check_connection_thread():
            try:
                connection = mysql.connector.connect(
                    host='192.168.1.220',
                    database='ztest',
                    user='test',
                    password='123',
                    charset='utf8'
                )
                if connection.is_connected():
                    print("Connected to database")  # Debugging line
                    Clock.schedule_once(lambda _: update_status_label("Connected to database", [0, 1, 0, 1]))
                else:
                    print("Not connected to database")  # Debugging line
                    Clock.schedule_once(lambda _: update_status_label("Not connected to database", [1, 0, 0, 1]))
            except Error as e:
                print(f"Error: {e}")  # Debugging line
                Clock.schedule_once(lambda _: update_status_label(f"Error: {e}", [1, 0, 0, 1]))
            finally:
                if connection and connection.is_connected():
                    connection.close()
    
            # Run the connection check in a separate thread to avoid blocking the main thread
        threading.Thread(target=check_connection_thread).start()

class MainApp(MDApp):
    def build(self):
        Builder.load_file('style.kv')
        return ReportExportWidget()

if __name__ == '__main__':
    MainApp().run()

